#!/usr/bin/env python3
"""Create comprehensive FY26 Revised Budget Excel with all class breakdowns, charts, and compensation."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize workbook
wb = Workbook()

# Style definitions
blue_font = Font(color="0000FF")  # Input cells
bold_font = Font(bold=True)
header_font = Font(bold=True, color="FFFFFF")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
header_fill = PatternFill("solid", fgColor="1C2833")
green_fill = PatternFill("solid", fgColor="27AE60")
light_gray_fill = PatternFill("solid", fgColor="F4F6F6")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
currency_format = '$#,##0.00;($#,##0.00);"-"'
percent_format = '0.0%;(0.0%);"-"'

# ============================================================================
# SHEET 1: FY26 Budget Highest Level
# ============================================================================
ws1 = wb.active
ws1.title = "FY26 Budget Highest Level"

# Header
ws1['A1'] = "Blue Drop"
ws1['A1'].font = Font(bold=True, size=16)
ws1['A2'] = "Budget FY26"
ws1['A2'].font = Font(bold=True, size=14)
ws1['A3'] = "October 2025 - September 2026"
ws1['A3'].font = Font(italic=True)

# Revenue section
ws1['A5'] = "Revenue"
ws1['B5'] = "Total"
ws1['C5'] = "Notes"
for cell in ['A5', 'B5', 'C5']:
    ws1[cell].font = header_font
    ws1[cell].fill = header_fill

revenue_items = [
    ("Bloom Sales", 425000, "Bloom Sales"),
    ("Bloom Deliveries", 125000, "Payment for deliveries from Bloom customers"),
    ("Bloom Hauling - DC Water", 1501808.37, "DC reimbursing for customer deliveries & Kiser"),
    ("Bloom - DC Water Marketing Fee", 1054130, "Includes Admin, Marketing, Sales Staff"),
    ("Land App/Storage", 749000, "Kiser and AmerEQ reimbursement"),
    ("Event Rental Revenue", 696000, "Goal for 52 events"),
    ("Products and IP Revenue", 340929, "Based on ARA's annual write ups"),
    ("Renewable Energy Income", 4360650, "Based on Chris's estimate"),
    ("Solar REC Sales (NEW)", 500000, "New solar REC revenue stream"),
    ("Cell Towers", 179971.86, "Based on current agreements"),
    ("Merchandise", 2750, "Blue Drop Shop (Books, Mugs, Shirts)"),
    ("Interest and Other", 290000, "Interest Rates and etc."),
]

row = 6
for name, amount, note in revenue_items:
    ws1.cell(row=row, column=1, value=name)
    ws1.cell(row=row, column=2, value=amount).font = blue_font
    ws1.cell(row=row, column=2).number_format = currency_format
    ws1.cell(row=row, column=3, value=note)
    row += 1

# Total Revenue
ws1.cell(row=row, column=1, value="Total Revenue").font = bold_font
ws1.cell(row=row, column=2, value=f"=SUM(B6:B{row-1})")
ws1.cell(row=row, column=2).number_format = currency_format
ws1.cell(row=row, column=2).font = bold_font
total_rev_row = row

row += 2

# Operating Expenditures
ws1.cell(row=row, column=1, value="Operating Expenditures").font = header_font
ws1.cell(row=row, column=1).fill = header_fill
ws1.cell(row=row, column=2, value="Total").font = header_font
ws1.cell(row=row, column=2).fill = header_fill
ws1.cell(row=row, column=3, value="Notes").font = header_font
ws1.cell(row=row, column=3).fill = header_fill
row += 1

expense_items = [
    ("Cost of Goods Sold", 2929949.60, "Hauling costs, Blending material, REC brokers, IP royalty payouts"),
    ("Compensation", 1085150.63, "See Compensation Math tab"),
    ("Professional Fees", 1044825.50, "Legal, Audit, Accounting, IT, HR, Consulting"),
    ("Travel", 11800, "Transportation, Lodging, Per Diem, Mileage"),
    ("Administration", 367823.78, "Office, Utilities, Insurance, Marketing"),
    ("Tax License & Insurance", 43925, "Business licenses and insurance"),
    ("Marketing", 61000, "Marketing and promotional expenses"),
    ("Interest and Fees", 2000, "Bank fees and interest"),
]

exp_start = row
for name, amount, note in expense_items:
    ws1.cell(row=row, column=1, value=name)
    ws1.cell(row=row, column=2, value=amount).font = blue_font
    ws1.cell(row=row, column=2).number_format = currency_format
    ws1.cell(row=row, column=3, value=note)
    row += 1

# Total Operating Expenditures
ws1.cell(row=row, column=1, value="Total Operation Expenditures").font = bold_font
ws1.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
ws1.cell(row=row, column=2).number_format = currency_format
ws1.cell(row=row, column=2).font = bold_font
total_exp_row = row

row += 2

# Net Profit
ws1.cell(row=row, column=1, value="Net Profit").font = Font(bold=True, size=14)
ws1.cell(row=row, column=2, value=f"=B{total_rev_row}-B{total_exp_row}")
ws1.cell(row=row, column=2).number_format = currency_format
ws1.cell(row=row, column=2).font = Font(bold=True, size=14, color="27AE60")

row += 2

# Capital Expenditures
ws1.cell(row=row, column=1, value="Total Capital Expenditures").font = bold_font
ws1.cell(row=row, column=2, value=500000).font = blue_font
ws1.cell(row=row, column=2).number_format = currency_format
ws1.cell(row=row, column=3, value="Purchase of land and capital improvements")

# Column widths
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 55

# ============================================================================
# SHEET 2: Compensation Math
# ============================================================================
ws2 = wb.create_sheet("Compensation Math")

# Allocation header
ws2['A1'] = "Staff"
ws2['B1'] = "Admin"
ws2['C1'] = "Events"
ws2['D1'] = "IP & Products"
ws2['E1'] = "Bloom Marketing"
ws2['F1'] = "Bloom Sales"
ws2['G1'] = "Bloom Admin"
ws2['H1'] = "Total"
for col in range(1, 9):
    ws2.cell(row=1, column=col).font = header_font
    ws2.cell(row=1, column=col).fill = header_fill

# Staff allocations
staff_allocations = [
    ("Alleyne, Victoria", 0, 0, 0, 0, 1, 0),
    ("EVENTS VACANT", 0, 1, 0, 0, 0, 0),
    ("Kiser, Holly", 0, 0, 0, 0, 1, 0),
    ("Lawrence, Adam", 0.5, 0, 0.25, 0, 0, 0.25),
    ("Thompson, April", 0, 0, 0, 0.4, 0.5, 0.1),
    ("Graham, Lovesha", 0.5, 0.25, 0, 0, 0, 0.25),
    ("Buddle, Raul", 0, 1, 0, 0, 0, 0),
    ("Klose, Cooper", 0, 0, 0, 0, 1, 0),
]

row = 2
for staff in staff_allocations:
    ws2.cell(row=row, column=1, value=staff[0])
    for i, val in enumerate(staff[1:], 2):
        if val > 0:
            ws2.cell(row=row, column=i, value=val).font = blue_font
    ws2.cell(row=row, column=8, value=f"=SUM(B{row}:G{row})")
    row += 1

row += 2

# Cost section header
ws2.cell(row=row, column=1, value="Cost").font = bold_font
row += 1
ws2.cell(row=row, column=1, value="Staff").font = header_font
ws2.cell(row=row, column=1).fill = header_fill
cost_headers = ["Current Base", "COL", "Performance Raise", "Proposed Base", "Bonus", "Proposed Salary", "Taxes", "Contributions", "Total"]
for i, h in enumerate(cost_headers, 2):
    ws2.cell(row=row, column=i, value=h).font = header_font
    ws2.cell(row=row, column=i).fill = header_fill
row += 1

# Staff costs
staff_costs = [
    ("Alleyne, Victoria", 77263.23, 0.03, 0, 30600, 8859.87, 23114.28),
    ("VACANT", 0, 0, 0, 0, 0, 0),
    ("Kiser, Holly", 66225.62, 0.03, 0, 40800, 7669.57, 34580.64),
    ("Lawrence, Adam", 82631.79, 0.03, 0, 0, 6818.85, 1262.52),
    ("Thompson, April", 133636.32, 0.03, 0, 15000, 11858.72, 26059.08),
    ("Graham, Lovesha", 141086, 0.03, 0, 15000, 13038.70, 36537.59),
    ("Buddle, Raul", 79950, 0.03, 0, 6000, 6454.73, 12179.52),
    ("Klose, Cooper", 68000, 0.015, 0, 0, 7982.64, 17000),
]

cost_start = row
for staff in staff_costs:
    ws2.cell(row=row, column=1, value=staff[0])
    ws2.cell(row=row, column=2, value=staff[1]).font = blue_font
    ws2.cell(row=row, column=2).number_format = currency_format
    ws2.cell(row=row, column=3, value=staff[2]).font = blue_font
    ws2.cell(row=row, column=3).number_format = percent_format
    ws2.cell(row=row, column=4, value=staff[3]).font = blue_font
    ws2.cell(row=row, column=4).number_format = percent_format
    # Proposed Base = Current * (1 + COL + Perf)
    ws2.cell(row=row, column=5, value=f"=B{row}*(1+C{row}+D{row})")
    ws2.cell(row=row, column=5).number_format = currency_format
    ws2.cell(row=row, column=6, value=staff[4]).font = blue_font
    ws2.cell(row=row, column=6).number_format = currency_format
    # Proposed Salary = Base + Bonus
    ws2.cell(row=row, column=7, value=f"=E{row}+F{row}")
    ws2.cell(row=row, column=7).number_format = currency_format
    ws2.cell(row=row, column=8, value=staff[5]).font = blue_font
    ws2.cell(row=row, column=8).number_format = currency_format
    ws2.cell(row=row, column=9, value=staff[6]).font = blue_font
    ws2.cell(row=row, column=9).number_format = currency_format
    # Total = Salary + Taxes + Contributions
    ws2.cell(row=row, column=10, value=f"=G{row}+H{row}+I{row}")
    ws2.cell(row=row, column=10).number_format = currency_format
    row += 1

# Total row
ws2.cell(row=row, column=1, value="Total").font = bold_font
for col in [2, 5, 6, 7, 8, 9, 10]:
    ws2.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{cost_start}:{get_column_letter(col)}{row-1})")
    ws2.cell(row=row, column=col).number_format = currency_format
    ws2.cell(row=row, column=col).font = bold_font

row += 3

# Cost Per Program section
ws2.cell(row=row, column=1, value="Cost Per Program").font = Font(bold=True, size=12)
row += 1
ws2.cell(row=row, column=1, value="Staff").font = header_font
ws2.cell(row=row, column=1).fill = header_fill
program_headers = ["Admin", "Events", "IP & Products", "Bloom Marketing", "Bloom Sales", "Bloom Admin", "Total"]
for i, h in enumerate(program_headers, 2):
    ws2.cell(row=row, column=i, value=h).font = header_font
    ws2.cell(row=row, column=i).fill = header_fill
row += 1

# Cost per program (calculated based on allocations and total costs)
program_costs = [
    ("Alleyne, Victoria", [0, 0, 0, 0, 142155.28, 0]),
    ("EVENTS VACANT", [0, 0, 0, 0, 0, 0]),
    ("Kiser, Holly", [0, 0, 0, 0, 151262.60, 0]),
    ("Lawrence, Adam", [46596.06, 0, 23298.03, 0, 0, 23298.03]),
    ("Thompson, April", [0, 0, 0, 76225.28, 95281.61, 19056.32]),
    ("Graham, Lovesha", [104947.44, 52473.72, 0, 0, 0, 52473.72]),
    ("Buddle, Raul", [0, 106982.75, 0, 0, 0, 0]),
    ("Klose, Cooper", [0, 0, 0, 0, 94002.64, 0]),
]

prog_start = row
for staff, costs in program_costs:
    ws2.cell(row=row, column=1, value=staff)
    for i, val in enumerate(costs, 2):
        ws2.cell(row=row, column=i, value=val).number_format = currency_format
    ws2.cell(row=row, column=8, value=f"=SUM(B{row}:G{row})")
    ws2.cell(row=row, column=8).number_format = currency_format
    row += 1

# Program totals
ws2.cell(row=row, column=1, value="Total").font = bold_font
for col in range(2, 9):
    ws2.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{prog_start}:{get_column_letter(col)}{row-1})")
    ws2.cell(row=row, column=col).number_format = currency_format
    ws2.cell(row=row, column=col).font = bold_font

for col in range(1, 11):
    ws2.column_dimensions[get_column_letter(col)].width = 16

# ============================================================================
# SHEET 3: FY26 Bloom PNL
# ============================================================================
ws3 = wb.create_sheet("FY26 Bloom PNL")

ws3['A1'] = "INCOME"
ws3['B1'] = "Bloom Marketing"
ws3['C1'] = "Bloom Sales"
ws3['D1'] = "TOTAL"
for cell in ['A1', 'B1', 'C1', 'D1']:
    ws3[cell].font = header_font
    ws3[cell].fill = header_fill

bloom_income = [
    ("4010 Bloom Sales", 0, 425000),
    ("4012 Bloom Deliveries", 0, 125000),
    ("4030 Bloom Hauling - DC Water", 1501808.37, 0),
    ("4031 Land App - DC Water reimbursement", 749000, 0),
    ("4040 Bloom - DC Water Marketing Fee", 1054130, 0),
]

row = 2
for name, mkt, sales in bloom_income:
    ws3.cell(row=row, column=1, value=name)
    ws3.cell(row=row, column=2, value=mkt).font = blue_font if mkt > 0 else Font()
    ws3.cell(row=row, column=2).number_format = currency_format
    ws3.cell(row=row, column=3, value=sales).font = blue_font if sales > 0 else Font()
    ws3.cell(row=row, column=3).number_format = currency_format
    ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
    ws3.cell(row=row, column=4).number_format = currency_format
    row += 1

# Total Income
ws3.cell(row=row, column=1, value="Total Income").font = bold_font
ws3.cell(row=row, column=2, value=f"=SUM(B2:B{row-1})")
ws3.cell(row=row, column=2).number_format = currency_format
ws3.cell(row=row, column=2).font = bold_font
ws3.cell(row=row, column=3, value=f"=SUM(C2:C{row-1})")
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=3).font = bold_font
ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
ws3.cell(row=row, column=4).number_format = currency_format
ws3.cell(row=row, column=4).font = bold_font
inc_row = row

row += 2

# COGS
ws3.cell(row=row, column=1, value="Cost of Goods Sold").font = header_font
ws3.cell(row=row, column=1).fill = header_fill
ws3.merge_cells(f'A{row}:D{row}')
row += 1

bloom_cogs = [
    ("5105 Bloom Hauling", 2176157.32, 0),
    ("5106 Blending Material", 0, 125062),
    ("5125 Subcontractors - COS", 7000, 0),
]

cogs_start = row
for name, mkt, sales in bloom_cogs:
    ws3.cell(row=row, column=1, value=name)
    ws3.cell(row=row, column=2, value=mkt).font = blue_font if mkt > 0 else Font()
    ws3.cell(row=row, column=2).number_format = currency_format
    ws3.cell(row=row, column=3, value=sales).font = blue_font if sales > 0 else Font()
    ws3.cell(row=row, column=3).number_format = currency_format
    ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
    ws3.cell(row=row, column=4).number_format = currency_format
    row += 1

ws3.cell(row=row, column=1, value="Total Cost of Goods Sold").font = bold_font
ws3.cell(row=row, column=2, value=f"=SUM(B{cogs_start}:B{row-1})")
ws3.cell(row=row, column=2).number_format = currency_format
ws3.cell(row=row, column=2).font = bold_font
ws3.cell(row=row, column=3, value=f"=SUM(C{cogs_start}:C{row-1})")
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=3).font = bold_font
ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
ws3.cell(row=row, column=4).number_format = currency_format
ws3.cell(row=row, column=4).font = bold_font
cogs_row = row

row += 1

# Gross Profit
ws3.cell(row=row, column=1, value="Gross Profit").font = Font(bold=True, color="27AE60")
ws3.cell(row=row, column=2, value=f"=B{inc_row}-B{cogs_row}")
ws3.cell(row=row, column=2).number_format = currency_format
ws3.cell(row=row, column=3, value=f"=C{inc_row}-C{cogs_row}")
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4, value=f"=D{inc_row}-D{cogs_row}")
ws3.cell(row=row, column=4).number_format = currency_format
ws3.cell(row=row, column=4).font = Font(bold=True, color="27AE60")
gp_row = row

row += 2

# Expenses
ws3.cell(row=row, column=1, value="Expenses").font = header_font
ws3.cell(row=row, column=1).fill = header_fill
ws3.merge_cells(f'A{row}:D{row}')
row += 1

bloom_expenses = [
    ("5010 Salary", 76225.28, 482702.12),
    ("5020 Fringe", 62188.21, 0),
    ("5030 Payroll Taxes", 20626.71, 0),
    ("5050 Bonus", 85817, 0),
    ("6001 Legal Fees", 0, 10000),
    ("6005 Human Resources Fees", 1500, 2750),
    ("6006 Operations Consulting", 85000, 0),
    ("6008 Land App/Storage", 749000, 0),
    ("6009 Other Professional Fees", 30000, 0),
]

exp_start = row
for name, mkt, sales in bloom_expenses:
    ws3.cell(row=row, column=1, value=name)
    ws3.cell(row=row, column=2, value=mkt).font = blue_font if mkt > 0 else Font()
    ws3.cell(row=row, column=2).number_format = currency_format
    ws3.cell(row=row, column=3, value=sales).font = blue_font if sales > 0 else Font()
    ws3.cell(row=row, column=3).number_format = currency_format
    ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
    ws3.cell(row=row, column=4).number_format = currency_format
    row += 1

ws3.cell(row=row, column=1, value="Total Expenses").font = bold_font
ws3.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
ws3.cell(row=row, column=2).number_format = currency_format
ws3.cell(row=row, column=2).font = bold_font
ws3.cell(row=row, column=3, value=f"=SUM(C{exp_start}:C{row-1})")
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=3).font = bold_font
ws3.cell(row=row, column=4, value=f"=B{row}+C{row}")
ws3.cell(row=row, column=4).number_format = currency_format
ws3.cell(row=row, column=4).font = bold_font
exp_tot_row = row

row += 2

# Net Income
ws3.cell(row=row, column=1, value="Net Income").font = Font(bold=True, size=12)
ws3.cell(row=row, column=2, value=f"=B{gp_row}-B{exp_tot_row}")
ws3.cell(row=row, column=2).number_format = currency_format
ws3.cell(row=row, column=3, value=f"=C{gp_row}-C{exp_tot_row}")
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4, value=f"=D{gp_row}-D{exp_tot_row}")
ws3.cell(row=row, column=4).number_format = currency_format
ws3.cell(row=row, column=4).font = Font(bold=True, size=12)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18

# ============================================================================
# SHEET 4: FY26 Events PNL
# ============================================================================
ws4 = wb.create_sheet("FY26 Events PNL")

ws4['A1'] = "Events PNL"
ws4['A1'].font = Font(bold=True, size=14)

ws4['A3'] = "Income"
ws4['B3'] = "Amount"
for cell in ['A3', 'B3']:
    ws4[cell].font = header_font
    ws4[cell].fill = header_fill

events_income = [
    ("4043 Event Rental Revenue", 696000),
]

row = 4
for name, amt in events_income:
    ws4.cell(row=row, column=1, value=name)
    ws4.cell(row=row, column=2, value=amt).font = blue_font
    ws4.cell(row=row, column=2).number_format = currency_format
    row += 1

ws4.cell(row=row, column=1, value="Total Income").font = bold_font
ws4.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})")
ws4.cell(row=row, column=2).number_format = currency_format
ws4.cell(row=row, column=2).font = bold_font
ev_inc_row = row

row += 2

# COGS
ws4.cell(row=row, column=1, value="Cost of Goods Sold").font = header_font
ws4.cell(row=row, column=1).fill = header_fill
row += 1

events_cogs = [
    ("5125 Subcontractors - COS", 54885.60),
    ("5129 Security & Valet", 35000),
]

cogs_start = row
for name, amt in events_cogs:
    ws4.cell(row=row, column=1, value=name)
    ws4.cell(row=row, column=2, value=amt).font = blue_font
    ws4.cell(row=row, column=2).number_format = currency_format
    row += 1

ws4.cell(row=row, column=1, value="Total COGS").font = bold_font
ws4.cell(row=row, column=2, value=f"=SUM(B{cogs_start}:B{row-1})")
ws4.cell(row=row, column=2).number_format = currency_format
ws4.cell(row=row, column=2).font = bold_font
ev_cogs_row = row

row += 1

ws4.cell(row=row, column=1, value="Gross Profit").font = Font(bold=True, color="27AE60")
ws4.cell(row=row, column=2, value=f"=B{ev_inc_row}-B{ev_cogs_row}")
ws4.cell(row=row, column=2).number_format = currency_format
ev_gp_row = row

row += 2

# Expenses
ws4.cell(row=row, column=1, value="Expenses").font = header_font
ws4.cell(row=row, column=1).fill = header_fill
row += 1

events_expenses = [
    ("5010 Salary", 82348.50),
    ("5020 Fringe", 13000),
    ("5030 Payroll Taxes", 7700),
    ("5050 Bonus", 6000),
    ("6004 IT Services", 11500),
    ("6005 Human Resources Fees", 3204),
    ("6009 Other Professional Fees (Valet & Security)", 79871.50),
    ("6101 Transportation", 500),
    ("6102 Lodging", 1500),
    ("6103 Per Diem", 500),
    ("6104 Mileage", 500),
    ("6202 Office Cleaning & Maintenance", 46500),
]

exp_start = row
for name, amt in events_expenses:
    ws4.cell(row=row, column=1, value=name)
    ws4.cell(row=row, column=2, value=amt).font = blue_font
    ws4.cell(row=row, column=2).number_format = currency_format
    row += 1

ws4.cell(row=row, column=1, value="Total Expenses").font = bold_font
ws4.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
ws4.cell(row=row, column=2).number_format = currency_format
ws4.cell(row=row, column=2).font = bold_font
ev_exp_row = row

row += 2

ws4.cell(row=row, column=1, value="Net Income").font = Font(bold=True, size=12)
ws4.cell(row=row, column=2, value=f"=B{ev_gp_row}-B{ev_exp_row}")
ws4.cell(row=row, column=2).number_format = currency_format
ws4.cell(row=row, column=2).font = Font(bold=True, size=12)

# Event Goals sidebar
ws4['E3'] = "FY26 Event Goals"
ws4['E3'].font = Font(bold=True, size=12)
ws4['E4'] = "Event Type"
ws4['F4'] = "# of Events"
ws4['G4'] = "Gross Revenue"
for cell in ['E4', 'F4', 'G4']:
    ws4[cell].font = header_font
    ws4[cell].fill = header_fill

event_goals = [
    ("Corporate meeting", 10, 145000),
    ("Free Community Event", 12, 138021),
    ("Individual Event", 10, 176940),
    ("Non-profit", 5, 76697),
    ("Government", 5, 48470),
    ("Wedding/Prom", 10, 110749),
]

row = 5
for name, num, rev in event_goals:
    ws4.cell(row=row, column=5, value=name)
    ws4.cell(row=row, column=6, value=num).font = blue_font
    ws4.cell(row=row, column=7, value=rev).font = blue_font
    ws4.cell(row=row, column=7).number_format = currency_format
    row += 1

ws4.cell(row=row, column=5, value="Total").font = bold_font
ws4.cell(row=row, column=6, value=f"=SUM(F5:F{row-1})")
ws4.cell(row=row, column=6).font = bold_font
ws4.cell(row=row, column=7, value=f"=SUM(G5:G{row-1})")
ws4.cell(row=row, column=7).number_format = currency_format
ws4.cell(row=row, column=7).font = bold_font

ws4.column_dimensions['A'].width = 45
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['E'].width = 22
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 15

# ============================================================================
# SHEET 5: FY26 IP PNL
# ============================================================================
ws5 = wb.create_sheet("FY26 IP PNL")

ws5['A1'] = "Income"
ws5['B1'] = "Amount"
ws5['C1'] = "Notes"
for cell in ['A1', 'B1', 'C1']:
    ws5[cell].font = header_font
    ws5[cell].fill = header_fill

ip_income = [
    ("4042 Products and IP Revenue", 340929, "Based on ARA projections"),
]

row = 2
for name, amt, note in ip_income:
    ws5.cell(row=row, column=1, value=name)
    ws5.cell(row=row, column=2, value=amt).font = blue_font
    ws5.cell(row=row, column=2).number_format = currency_format
    ws5.cell(row=row, column=3, value=note)
    row += 1

ws5.cell(row=row, column=1, value="Total Income").font = bold_font
ws5.cell(row=row, column=2, value=f"=SUM(B2:B{row-1})")
ws5.cell(row=row, column=2).number_format = currency_format
ws5.cell(row=row, column=2).font = bold_font
ip_inc_row = row

row += 2

ws5.cell(row=row, column=1, value="Cost of Goods Sold").font = header_font
ws5.cell(row=row, column=1).fill = header_fill
row += 1

ip_cogs = [
    ("5115 Other Costs - COS (Royalty payout)", 132960, "Royalty pay out"),
]

cogs_start = row
for name, amt, note in ip_cogs:
    ws5.cell(row=row, column=1, value=name)
    ws5.cell(row=row, column=2, value=amt).font = blue_font
    ws5.cell(row=row, column=2).number_format = currency_format
    ws5.cell(row=row, column=3, value=note)
    row += 1

ws5.cell(row=row, column=1, value="Total COGS").font = bold_font
ws5.cell(row=row, column=2, value=f"=SUM(B{cogs_start}:B{row-1})")
ws5.cell(row=row, column=2).number_format = currency_format
ws5.cell(row=row, column=2).font = bold_font
ip_cogs_row = row

row += 1

ws5.cell(row=row, column=1, value="Gross Profit").font = Font(bold=True, color="27AE60")
ws5.cell(row=row, column=2, value=f"=B{ip_inc_row}-B{ip_cogs_row}")
ws5.cell(row=row, column=2).number_format = currency_format
ip_gp_row = row

row += 2

ws5.cell(row=row, column=1, value="Expenses").font = header_font
ws5.cell(row=row, column=1).fill = header_fill
row += 1

ip_expenses = [
    ("5000 Compensation (allocated)", 23298.03, "25% of Adam Lawrence"),
]

exp_start = row
for name, amt, note in ip_expenses:
    ws5.cell(row=row, column=1, value=name)
    ws5.cell(row=row, column=2, value=amt).font = blue_font
    ws5.cell(row=row, column=2).number_format = currency_format
    ws5.cell(row=row, column=3, value=note)
    row += 1

ws5.cell(row=row, column=1, value="Total Expenses").font = bold_font
ws5.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
ws5.cell(row=row, column=2).number_format = currency_format
ws5.cell(row=row, column=2).font = bold_font
ip_exp_row = row

row += 2

ws5.cell(row=row, column=1, value="Net Income").font = Font(bold=True, size=12)
ws5.cell(row=row, column=2, value=f"=B{ip_gp_row}-B{ip_exp_row}")
ws5.cell(row=row, column=2).number_format = currency_format
ws5.cell(row=row, column=2).font = Font(bold=True, size=12)

# IP Revenue breakdown
row += 3
ws5.cell(row=row, column=1, value="IP Revenue Breakdown").font = Font(bold=True, size=12)
row += 1
ws5.cell(row=row, column=1, value="IP").font = header_font
ws5.cell(row=row, column=1).fill = header_fill
ws5.cell(row=row, column=2, value="Projected Revenue").font = header_font
ws5.cell(row=row, column=2).fill = header_fill
ws5.cell(row=row, column=3, value="Notes").font = header_font
ws5.cell(row=row, column=3).fill = header_fill
row += 1

ip_breakdown = [
    ("Ovivo", 100000, "Minimum Contract Amount"),
    ("inDense", 93000, "25% royalty on ARA projected €341k"),
    ("DEMON", 156000, "50% royalty on ARA projected $315k"),
    ("LayerMark", 0, "TBD"),
]

ip_start = row
for name, amt, note in ip_breakdown:
    ws5.cell(row=row, column=1, value=name)
    ws5.cell(row=row, column=2, value=amt).font = blue_font
    ws5.cell(row=row, column=2).number_format = currency_format
    ws5.cell(row=row, column=3, value=note)
    row += 1

ws5.cell(row=row, column=1, value="Total").font = bold_font
ws5.cell(row=row, column=2, value=f"=SUM(B{ip_start}:B{row-1})")
ws5.cell(row=row, column=2).number_format = currency_format
ws5.cell(row=row, column=2).font = bold_font

ws5.column_dimensions['A'].width = 40
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 40

# ============================================================================
# SHEET 6: FY26 RECs PNL
# ============================================================================
ws6 = wb.create_sheet("FY26 RECs PNL")

ws6['A1'] = "Renewable Energy Credits (RECs) PNL"
ws6['A1'].font = Font(bold=True, size=14)

ws6['A3'] = "Income"
ws6['B3'] = "Amount"
ws6['C3'] = "Notes"
for cell in ['A3', 'B3', 'C3']:
    ws6[cell].font = header_font
    ws6[cell].fill = header_fill

recs_income = [
    ("4044 REC Sales - Tier 1", 4100782, "Based on Chris's estimate"),
    ("4044 REC Sales - HEX", 256416, "Hex credits"),
    ("4044 Solar RECs (NEW)", 500000, "New solar REC stream"),
    ("4065 REC Admin Fee", 21803, "Admin fee on REC sales"),
]

row = 4
for name, amt, note in recs_income:
    ws6.cell(row=row, column=1, value=name)
    ws6.cell(row=row, column=2, value=amt).font = blue_font
    ws6.cell(row=row, column=2).number_format = currency_format
    ws6.cell(row=row, column=3, value=note)
    row += 1

ws6.cell(row=row, column=1, value="Total Income").font = bold_font
ws6.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})")
ws6.cell(row=row, column=2).number_format = currency_format
ws6.cell(row=row, column=2).font = bold_font
rec_inc_row = row

row += 2

ws6.cell(row=row, column=1, value="Cost of Goods Sold").font = header_font
ws6.cell(row=row, column=1).fill = header_fill
row += 1

recs_cogs = [
    ("5120 REC Broker Fees", 217803, "Broker commission on REC sales"),
]

cogs_start = row
for name, amt, note in recs_cogs:
    ws6.cell(row=row, column=1, value=name)
    ws6.cell(row=row, column=2, value=amt).font = blue_font
    ws6.cell(row=row, column=2).number_format = currency_format
    ws6.cell(row=row, column=3, value=note)
    row += 1

ws6.cell(row=row, column=1, value="Total COGS").font = bold_font
ws6.cell(row=row, column=2, value=f"=SUM(B{cogs_start}:B{row-1})")
ws6.cell(row=row, column=2).number_format = currency_format
ws6.cell(row=row, column=2).font = bold_font
rec_cogs_row = row

row += 1

ws6.cell(row=row, column=1, value="Gross Profit").font = Font(bold=True, color="27AE60")
ws6.cell(row=row, column=2, value=f"=B{rec_inc_row}-B{rec_cogs_row}")
ws6.cell(row=row, column=2).number_format = currency_format
rec_gp_row = row

row += 2

ws6.cell(row=row, column=1, value="Expenses").font = header_font
ws6.cell(row=row, column=1).fill = header_fill
row += 1

ws6.cell(row=row, column=1, value="No direct expenses allocated")

row += 2

ws6.cell(row=row, column=1, value="Net Income").font = Font(bold=True, size=12)
ws6.cell(row=row, column=2, value=f"=B{rec_gp_row}")
ws6.cell(row=row, column=2).number_format = currency_format
ws6.cell(row=row, column=2).font = Font(bold=True, size=12)

# REC pricing table
row += 3
ws6.cell(row=row, column=1, value="REC Pricing & Volume").font = Font(bold=True, size=12)
row += 1
ws6.cell(row=row, column=1, value="Type").font = header_font
ws6.cell(row=row, column=1).fill = header_fill
ws6.cell(row=row, column=2, value="Volume").font = header_font
ws6.cell(row=row, column=2).fill = header_fill
ws6.cell(row=row, column=3, value="DC Price").font = header_font
ws6.cell(row=row, column=3).fill = header_fill
ws6.cell(row=row, column=4, value="MD/PA Price").font = header_font
ws6.cell(row=row, column=4).fill = header_fill
ws6.cell(row=row, column=5, value="Revenue").font = header_font
ws6.cell(row=row, column=5).fill = header_fill
row += 1

rec_pricing = [
    ("Tier 1", 165700, 30, 34),
    ("HEX", 14325, 30, 34),
    ("SREC", 10, 440, 0),
]

rec_start = row
for name, vol, dc, mdpa in rec_pricing:
    ws6.cell(row=row, column=1, value=name)
    ws6.cell(row=row, column=2, value=vol).font = blue_font
    ws6.cell(row=row, column=3, value=dc).font = blue_font
    ws6.cell(row=row, column=3).number_format = '$#,##0'
    ws6.cell(row=row, column=4, value=mdpa).font = blue_font
    ws6.cell(row=row, column=4).number_format = '$#,##0'
    ws6.cell(row=row, column=5, value=f"=B{row}*((C{row}+D{row})/2)")
    ws6.cell(row=row, column=5).number_format = currency_format
    row += 1

ws6.cell(row=row, column=1, value="Total").font = bold_font
ws6.cell(row=row, column=5, value=f"=SUM(E{rec_start}:E{row-1})")
ws6.cell(row=row, column=5).number_format = currency_format
ws6.cell(row=row, column=5).font = bold_font

ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15

# ============================================================================
# SHEET 7: FY26 Admin PNL
# ============================================================================
ws7 = wb.create_sheet("FY26 Admin PNL")

ws7['A1'] = "Administration PNL"
ws7['A1'].font = Font(bold=True, size=14)

ws7['A3'] = "Income"
ws7['B3'] = "Amount"
ws7['C3'] = "Notes"
for cell in ['A3', 'B3', 'C3']:
    ws7[cell].font = header_font
    ws7[cell].fill = header_fill

admin_income = [
    ("4045 Cell Tower Revenue", 179971.86, "Based on current agreements"),
    ("4050 Interest", 290000, "Interest income from cash reserves"),
    ("4070 Web Store", 2750, "Blue Drop Shop"),
]

row = 4
for name, amt, note in admin_income:
    ws7.cell(row=row, column=1, value=name)
    ws7.cell(row=row, column=2, value=amt).font = blue_font
    ws7.cell(row=row, column=2).number_format = currency_format
    ws7.cell(row=row, column=3, value=note)
    row += 1

ws7.cell(row=row, column=1, value="Total Income").font = bold_font
ws7.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})")
ws7.cell(row=row, column=2).number_format = currency_format
ws7.cell(row=row, column=2).font = bold_font
admin_inc_row = row

row += 2

ws7.cell(row=row, column=1, value="Expenses").font = header_font
ws7.cell(row=row, column=1).fill = header_fill
row += 1

admin_expenses = [
    ("5010 Salary (allocated)", 151543.49, "Admin allocation of Lovesha & Adam"),
    ("5020 Fringe", 36537.59, "Benefits"),
    ("5030 Payroll Taxes", 13038.70, "Payroll taxes"),
    ("5050 Bonus", 15000, "Bonus allocation"),
    ("6001 Legal Fees", 75000, "Legal services"),
    ("6002 Audit Fees", 24000, "Annual audit"),
    ("6003 Accounting Fees", 55200, "Accounting services"),
    ("6004 IT Services", 58800, "IT infrastructure"),
    ("6005 Human Resources Fees", 30000, "HR services"),
    ("6009 Other Professional Fees", 25000, "Miscellaneous consulting"),
    ("6201 Bank Fees", 2000, "Banking fees"),
    ("6203 Utilities", 12000, "Office utilities"),
    ("6205 Office Supplies", 5000, "Supplies"),
    ("6207 Software Subscriptions", 24000, "Software licenses"),
    ("6301 D&O Insurance", 43925, "Directors & Officers insurance"),
    ("6401 Marketing & Promotion", 61000, "Marketing expenses"),
]

exp_start = row
for name, amt, note in admin_expenses:
    ws7.cell(row=row, column=1, value=name)
    ws7.cell(row=row, column=2, value=amt).font = blue_font
    ws7.cell(row=row, column=2).number_format = currency_format
    ws7.cell(row=row, column=3, value=note)
    row += 1

ws7.cell(row=row, column=1, value="Total Expenses").font = bold_font
ws7.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
ws7.cell(row=row, column=2).number_format = currency_format
ws7.cell(row=row, column=2).font = bold_font
admin_exp_row = row

row += 2

ws7.cell(row=row, column=1, value="Net Income").font = Font(bold=True, size=12)
ws7.cell(row=row, column=2, value=f"=B{admin_inc_row}-B{admin_exp_row}")
ws7.cell(row=row, column=2).number_format = currency_format
ws7.cell(row=row, column=2).font = Font(bold=True, size=12)

ws7.column_dimensions['A'].width = 35
ws7.column_dimensions['B'].width = 18
ws7.column_dimensions['C'].width = 35

# ============================================================================
# SHEET 8: Income Math
# ============================================================================
ws8 = wb.create_sheet("Income Math")

ws8['A1'] = "Products and IP"
ws8['A1'].font = Font(bold=True, size=12)
ws8['A2'] = "IP"
ws8['B2'] = "Projected Revenue"
ws8['C2'] = "Notes"
for cell in ['A2', 'B2', 'C2']:
    ws8[cell].font = header_font
    ws8[cell].fill = header_fill

ip_math = [
    ("Ovivo", 100000, "Minimum Contract Amount"),
    ("inDense", 93000, "25% royalty on ARA projected €341k"),
    ("DEMON", 156000, "50% royalty on ARA projected $315k"),
    ("LayerMark", 0, "TBD"),
]

row = 3
ip_start = row
for name, amt, note in ip_math:
    ws8.cell(row=row, column=1, value=name)
    ws8.cell(row=row, column=2, value=amt).font = blue_font
    ws8.cell(row=row, column=2).number_format = currency_format
    ws8.cell(row=row, column=3, value=note)
    row += 1

ws8.cell(row=row, column=1, value="Total").font = bold_font
ws8.cell(row=row, column=2, value=f"=SUM(B{ip_start}:B{row-1})")
ws8.cell(row=row, column=2).number_format = currency_format
ws8.cell(row=row, column=2).font = bold_font

row += 3

# Wendy's Store
ws8.cell(row=row, column=1, value="Wendy's Store").font = Font(bold=True, size=12)
row += 1
ws8.cell(row=row, column=1, value="Product").font = header_font
ws8.cell(row=row, column=1).fill = header_fill
ws8.cell(row=row, column=2, value="Amount Sold").font = header_font
ws8.cell(row=row, column=2).fill = header_fill
row += 1

store_items = [
    ("Book", 500),
    ("Book to Utilities", 2000),
    ("Store", 250),
]

store_start = row
for name, amt in store_items:
    ws8.cell(row=row, column=1, value=name)
    ws8.cell(row=row, column=2, value=amt).font = blue_font
    ws8.cell(row=row, column=2).number_format = currency_format
    row += 1

ws8.cell(row=row, column=1, value="Total").font = bold_font
ws8.cell(row=row, column=2, value=f"=SUM(B{store_start}:B{row-1})")
ws8.cell(row=row, column=2).number_format = currency_format
ws8.cell(row=row, column=2).font = bold_font

row += 3

# Cell Towers
ws8.cell(row=row, column=1, value="Cell Towers").font = Font(bold=True, size=12)
row += 1
ws8.cell(row=row, column=1, value="Location").font = header_font
ws8.cell(row=row, column=1).fill = header_fill
ws8.cell(row=row, column=2, value="Carrier").font = header_font
ws8.cell(row=row, column=2).fill = header_fill
ws8.cell(row=row, column=3, value="Annual Amount").font = header_font
ws8.cell(row=row, column=3).fill = header_fill
row += 1

cell_towers = [
    ("2500 25th St SE", "AT&T", 53240.48),
    ("2500 25th St SE", "Dish", 28538.47),
    ("5000 Overlook Ave SW", "Verizon", 0),
    ("5000 Overlook Ave SW", "New Cingular", 43892.91),
    ("5000 Overlook Ave SW", "T-Mobile-new", 54300),
]

tower_start = row
for loc, carrier, amt in cell_towers:
    ws8.cell(row=row, column=1, value=loc)
    ws8.cell(row=row, column=2, value=carrier)
    ws8.cell(row=row, column=3, value=amt).font = blue_font
    ws8.cell(row=row, column=3).number_format = currency_format
    row += 1

ws8.cell(row=row, column=1, value="Total").font = bold_font
ws8.cell(row=row, column=3, value=f"=SUM(C{tower_start}:C{row-1})")
ws8.cell(row=row, column=3).number_format = currency_format
ws8.cell(row=row, column=3).font = bold_font

ws8.column_dimensions['A'].width = 25
ws8.column_dimensions['B'].width = 20
ws8.column_dimensions['C'].width = 45

# ============================================================================
# SHEET 9: Dashboard with Charts
# ============================================================================
ws9 = wb.create_sheet("Dashboard")

ws9['A1'] = "FY26 Budget Dashboard"
ws9['A1'].font = Font(bold=True, size=16)

# Revenue by Category table for chart
ws9['A4'] = "Revenue by Category"
ws9['A4'].font = Font(bold=True, size=12)
ws9['A5'] = "Category"
ws9['B5'] = "Amount"
for cell in ['A5', 'B5']:
    ws9[cell].font = header_font
    ws9[cell].fill = header_fill

revenue_chart_data = [
    ("RECs", 4879001),
    ("Bloom Marketing", 3304938),
    ("Events", 696000),
    ("IP & Products", 340929),
    ("Cell Towers", 179972),
    ("Interest & Other", 290000),
    ("Store", 2750),
]

row = 6
for name, amt in revenue_chart_data:
    ws9.cell(row=row, column=1, value=name)
    ws9.cell(row=row, column=2, value=amt)
    ws9.cell(row=row, column=2).number_format = currency_format
    row += 1

# Create Pie Chart
pie = PieChart()
labels = Reference(ws9, min_col=1, min_row=6, max_row=12)
data = Reference(ws9, min_col=2, min_row=5, max_row=12)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "FY26 Revenue Mix"
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = False
pie.dataLabels.showVal = False
ws9.add_chart(pie, "D4")

# Net Income by Segment table
ws9['A16'] = "Net Income by Segment"
ws9['A16'].font = Font(bold=True, size=12)
ws9['A17'] = "Segment"
ws9['B17'] = "Net Income"
for cell in ['A17', 'B17']:
    ws9[cell].font = header_font
    ws9[cell].fill = header_fill

segment_ni = [
    ("RECs", 4661198),
    ("IP & Products", 184671),
    ("Admin/Other", -159322),
    ("Bloom Combined", 159781),
    ("Events", 127490),
]

row = 18
for name, amt in segment_ni:
    ws9.cell(row=row, column=1, value=name)
    ws9.cell(row=row, column=2, value=amt)
    ws9.cell(row=row, column=2).number_format = currency_format
    row += 1

# Create Bar Chart
bar = BarChart()
bar.type = "col"
bar.style = 10
bar.title = "Net Income by Segment"
bar.y_axis.title = "Net Income ($)"
bar_data = Reference(ws9, min_col=2, min_row=17, max_row=22)
bar_cats = Reference(ws9, min_col=1, min_row=18, max_row=22)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.shape = 4
ws9.add_chart(bar, "D18")

# Key Metrics
ws9['A25'] = "Key Metrics"
ws9['A25'].font = Font(bold=True, size=12)

metrics = [
    ("Total Revenue", "='FY26 Budget Highest Level'!B17"),
    ("Total Expenses", "='FY26 Budget Highest Level'!B27"),
    ("Net Profit", "='FY26 Budget Highest Level'!B29"),
    ("Profit Margin", "=B28/B26"),
    ("Total Compensation", "='Compensation Math'!J20"),
]

row = 26
for name, formula in metrics:
    ws9.cell(row=row, column=1, value=name)
    ws9.cell(row=row, column=2, value=formula)
    if "Margin" in name:
        ws9.cell(row=row, column=2).number_format = percent_format
    else:
        ws9.cell(row=row, column=2).number_format = currency_format
    row += 1

ws9.column_dimensions['A'].width = 20
ws9.column_dimensions['B'].width = 18

# ============================================================================
# SHEET 10: Review Process
# ============================================================================
ws10 = wb.create_sheet("Review Process")

ws10['A1'] = "Budget Review Process"
ws10['A1'].font = Font(bold=True, size=16)

ws10['A3'] = "Review Schedule"
ws10['A3'].font = Font(bold=True, size=12)

review_schedule = [
    ("Weekly", "Cash position review", "CFO"),
    ("Monthly", "Variance analysis vs budget", "CFO + Budget Owners"),
    ("Quarterly", "Board financial presentation", "CFO + CEO"),
    ("Annual", "Full budget revision", "Leadership Team"),
]

ws10['A4'] = "Frequency"
ws10['B4'] = "Activity"
ws10['C4'] = "Responsible"
for cell in ['A4', 'B4', 'C4']:
    ws10[cell].font = header_font
    ws10[cell].fill = header_fill

row = 5
for freq, activity, resp in review_schedule:
    ws10.cell(row=row, column=1, value=freq)
    ws10.cell(row=row, column=2, value=activity)
    ws10.cell(row=row, column=3, value=resp)
    row += 1

row += 2

ws10.cell(row=row, column=1, value="Variance Thresholds").font = Font(bold=True, size=12)
row += 1

thresholds = [
    ("< ±5%", "Monitor only", "No action required"),
    ("±5% to ±10%", "Investigation", "Document reason for variance"),
    ("±10% to ±20%", "Management review", "Present to leadership"),
    ("> ±20%", "Board notification", "Include in board report"),
]

ws10.cell(row=row, column=1, value="Variance").font = header_font
ws10.cell(row=row, column=1).fill = header_fill
ws10.cell(row=row, column=2, value="Action").font = header_font
ws10.cell(row=row, column=2).fill = header_fill
ws10.cell(row=row, column=3, value="Requirement").font = header_font
ws10.cell(row=row, column=3).fill = header_fill
row += 1

for var, action, req in thresholds:
    ws10.cell(row=row, column=1, value=var)
    ws10.cell(row=row, column=2, value=action)
    ws10.cell(row=row, column=3, value=req)
    row += 1

row += 2

ws10.cell(row=row, column=1, value="Budget Owners").font = Font(bold=True, size=12)
row += 1

owners = [
    ("RECs", "Chris / Adam"),
    ("Bloom Marketing", "April / Holly"),
    ("Bloom Sales", "Victoria / Cooper"),
    ("Events", "Raul / Lovesha"),
    ("IP & Products", "Adam"),
    ("Administration", "Lovesha"),
]

ws10.cell(row=row, column=1, value="Budget Area").font = header_font
ws10.cell(row=row, column=1).fill = header_fill
ws10.cell(row=row, column=2, value="Owner").font = header_font
ws10.cell(row=row, column=2).fill = header_fill
row += 1

for area, owner in owners:
    ws10.cell(row=row, column=1, value=area)
    ws10.cell(row=row, column=2, value=owner)
    row += 1

ws10.column_dimensions['A'].width = 20
ws10.column_dimensions['B'].width = 35
ws10.column_dimensions['C'].width = 30

# Save workbook
output_path = "/Users/loveg/Documents/SDO - Workspace/FY26_Revised_Budget_Board_Presentation.xlsx"
wb.save(output_path)
print(f"Created: {output_path}")
print("Sheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
