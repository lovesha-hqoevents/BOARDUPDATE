#!/usr/bin/env python3
"""Add YTD vs Budget comparison sheet to the FY26 Revised Budget."""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Load existing workbook
wb = load_workbook("/Users/loveg/Documents/SDO - Workspace/FY26_Revised_Budget_Board_Presentation.xlsx")

# Style definitions
blue_font = Font(color="0000FF")
bold_font = Font(bold=True)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1C2833")
green_fill = PatternFill("solid", fgColor="D5F5E3")
red_fill = PatternFill("solid", fgColor="FADBD8")
yellow_fill = PatternFill("solid", fgColor="FCF3CF")
currency_format = '$#,##0;($#,##0);"-"'
percent_format = '0.0%;(0.0%);"-"'

# ============================================================================
# Create YTD vs Budget Sheet (insert after FY26 Budget Highest Level)
# ============================================================================
ws = wb.create_sheet("YTD vs Budget", 1)

ws['A1'] = "FY26 YTD Performance vs Budget"
ws['A1'].font = Font(bold=True, size=16)
ws['A2'] = "As of January 19, 2026 (Q1 + 19 days of Q2)"
ws['A2'].font = Font(italic=True, size=11)

# Period info
ws['A4'] = "Budget Period:"
ws['B4'] = "Q1 (25% of Annual)"
ws['A5'] = "Actual Period:"
ws['B5'] = "Oct 1, 2025 - Jan 19, 2026"
ws['A6'] = "Days in FY26:"
ws['B6'] = 111
ws['C6'] = "of 365 (30.4%)"

row = 8

# Revenue Comparison Header
ws.cell(row=row, column=1, value="Revenue Category").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=2, value="FY26 Budget").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=3, value="Q1 Budget (25%)").font = header_font
ws.cell(row=row, column=3).fill = header_fill
ws.cell(row=row, column=4, value="YTD Actual").font = header_font
ws.cell(row=row, column=4).fill = header_fill
ws.cell(row=row, column=5, value="Variance $").font = header_font
ws.cell(row=row, column=5).fill = header_fill
ws.cell(row=row, column=6, value="% of Q1").font = header_font
ws.cell(row=row, column=6).fill = header_fill
ws.cell(row=row, column=7, value="Full Year Forecast").font = header_font
ws.cell(row=row, column=7).fill = header_fill
ws.cell(row=row, column=8, value="Status").font = header_font
ws.cell(row=row, column=8).fill = header_fill
row += 1

# YTD Actuals from P&L by Class (Oct 1, 2025 - Jan 19, 2026)
revenue_data = [
    # (Category, FY26 Budget, YTD Actual)
    ("RECs (Renewable Energy)", 4860650, 2518742),
    ("Bloom Marketing (DC Water Fees)", 3304938, 261185),
    ("Bloom Sales", 550000, 90769),
    ("DC Water IP/Licensing", 340929, 61500),
    ("Events", 696000, 43299),
    ("Cell Towers", 179972, 43182),
    ("Interest Income", 290000, 72048),
    ("Store/Merchandise", 2750, 0),
    ("Other Revenue", 0, 30232),
]

rev_start = row
for name, budget, actual in revenue_data:
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=budget).font = blue_font
    ws.cell(row=row, column=2).number_format = currency_format
    ws.cell(row=row, column=3, value=f"=B{row}*0.25")  # Q1 = 25%
    ws.cell(row=row, column=3).number_format = currency_format
    ws.cell(row=row, column=4, value=actual).font = blue_font
    ws.cell(row=row, column=4).number_format = currency_format
    ws.cell(row=row, column=5, value=f"=D{row}-C{row}")  # Variance
    ws.cell(row=row, column=5).number_format = currency_format
    ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")  # % of Q1
    ws.cell(row=row, column=6).number_format = percent_format
    ws.cell(row=row, column=7, value=f"=D{row}/$B$6*365")  # Annualized forecast
    ws.cell(row=row, column=7).number_format = currency_format
    # Status formula
    ws.cell(row=row, column=8, value=f'=IF(F{row}>=1,"ON TRACK",IF(F{row}>=0.8,"MONITOR","UNDER"))')
    row += 1

# Total Revenue
ws.cell(row=row, column=1, value="TOTAL REVENUE").font = bold_font
for col in [2, 3, 4, 5, 7]:
    ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{rev_start}:{get_column_letter(col)}{row-1})")
    ws.cell(row=row, column=col).number_format = currency_format
    ws.cell(row=row, column=col).font = bold_font
ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
ws.cell(row=row, column=6).number_format = percent_format
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=8, value=f'=IF(F{row}>=1,"ON TRACK",IF(F{row}>=0.8,"MONITOR","UNDER"))')
ws.cell(row=row, column=8).font = bold_font
rev_total_row = row

row += 2

# Expense Comparison Header
ws.cell(row=row, column=1, value="Expense Category").font = header_font
ws.cell(row=row, column=1).fill = header_fill
for col, header in enumerate(["FY26 Budget", "Q1 Budget (25%)", "YTD Actual", "Variance $", "% of Q1", "Full Year Forecast", "Status"], 2):
    ws.cell(row=row, column=col, value=header).font = header_font
    ws.cell(row=row, column=col).fill = header_fill
row += 1

# YTD Expense Actuals
expense_data = [
    # (Category, FY26 Budget, YTD Actual)
    ("Cost of Goods Sold", 2929950, 847286),
    ("Compensation", 1085151, 229734),
    ("Professional Fees", 1044826, 68570),
    ("Travel", 11800, 1250),
    ("Administration", 367824, 14449),
    ("Tax/License/Insurance", 43925, 2000),
    ("Marketing", 61000, 0),
    ("Interest & Fees", 2000, 0),
]

exp_start = row
for name, budget, actual in expense_data:
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=budget).font = blue_font
    ws.cell(row=row, column=2).number_format = currency_format
    ws.cell(row=row, column=3, value=f"=B{row}*0.25")
    ws.cell(row=row, column=3).number_format = currency_format
    ws.cell(row=row, column=4, value=actual).font = blue_font
    ws.cell(row=row, column=4).number_format = currency_format
    ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
    ws.cell(row=row, column=5).number_format = currency_format
    ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
    ws.cell(row=row, column=6).number_format = percent_format
    ws.cell(row=row, column=7, value=f"=D{row}/$B$6*365")
    ws.cell(row=row, column=7).number_format = currency_format
    # For expenses, under budget is favorable
    ws.cell(row=row, column=8, value=f'=IF(F{row}<=1,"FAVORABLE",IF(F{row}<=1.1,"MONITOR","OVER"))')
    row += 1

# Total Expenses
ws.cell(row=row, column=1, value="TOTAL EXPENSES").font = bold_font
for col in [2, 3, 4, 5, 7]:
    ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{exp_start}:{get_column_letter(col)}{row-1})")
    ws.cell(row=row, column=col).number_format = currency_format
    ws.cell(row=row, column=col).font = bold_font
ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
ws.cell(row=row, column=6).number_format = percent_format
ws.cell(row=row, column=6).font = bold_font
ws.cell(row=row, column=8, value=f'=IF(F{row}<=1,"FAVORABLE",IF(F{row}<=1.1,"MONITOR","OVER"))')
ws.cell(row=row, column=8).font = bold_font
exp_total_row = row

row += 2

# Net Income
ws.cell(row=row, column=1, value="NET INCOME").font = Font(bold=True, size=12)
ws.cell(row=row, column=2, value=f"=B{rev_total_row}-B{exp_total_row}")
ws.cell(row=row, column=2).number_format = currency_format
ws.cell(row=row, column=2).font = Font(bold=True, size=12)
ws.cell(row=row, column=3, value=f"=C{rev_total_row}-C{exp_total_row}")
ws.cell(row=row, column=3).number_format = currency_format
ws.cell(row=row, column=3).font = Font(bold=True, size=12)
ws.cell(row=row, column=4, value=f"=D{rev_total_row}-D{exp_total_row}")
ws.cell(row=row, column=4).number_format = currency_format
ws.cell(row=row, column=4).font = Font(bold=True, size=12, color="27AE60")
ws.cell(row=row, column=5, value=f"=E{rev_total_row}-E{exp_total_row}")
ws.cell(row=row, column=5).number_format = currency_format
ws.cell(row=row, column=5).font = Font(bold=True, size=12)
ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
ws.cell(row=row, column=6).number_format = percent_format
ws.cell(row=row, column=6).font = Font(bold=True, size=12)
ws.cell(row=row, column=7, value=f"=G{rev_total_row}-G{exp_total_row}")
ws.cell(row=row, column=7).number_format = currency_format
ws.cell(row=row, column=7).font = Font(bold=True, size=12)
ni_row = row

row += 3

# Segment Performance Summary
ws.cell(row=row, column=1, value="Segment Performance Summary").font = Font(bold=True, size=12)
row += 1
ws.cell(row=row, column=1, value="Segment").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=2, value="YTD Revenue").font = header_font
ws.cell(row=row, column=2).fill = header_fill
ws.cell(row=row, column=3, value="YTD Expenses").font = header_font
ws.cell(row=row, column=3).fill = header_fill
ws.cell(row=row, column=4, value="YTD Net Income").font = header_font
ws.cell(row=row, column=4).fill = header_fill
ws.cell(row=row, column=5, value="% of Total NI").font = header_font
ws.cell(row=row, column=5).fill = header_fill
ws.cell(row=row, column=6, value="Status").font = header_font
ws.cell(row=row, column=6).fill = header_fill
row += 1

# Segment data from YTD actuals
segment_data = [
    ("RECs", 2518742, 126094, "STRONG"),
    ("Bloom Marketing", 261185, 847351, "LOSS"),
    ("Bloom Sales", 90769, 39536, "PROFIT"),
    ("Events", 43299, 94123, "LOSS"),
    ("IP & Products", 61500, 10000, "PROFIT"),
    ("Administration", 145230, 46185, "PROFIT"),
]

seg_start = row
for name, rev, exp, status in segment_data:
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=rev).font = blue_font
    ws.cell(row=row, column=2).number_format = currency_format
    ws.cell(row=row, column=3, value=exp).font = blue_font
    ws.cell(row=row, column=3).number_format = currency_format
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=4).number_format = currency_format
    ws.cell(row=row, column=5, value=f"=IF($D${ni_row}=0,0,D{row}/$D${ni_row})")
    ws.cell(row=row, column=5).number_format = percent_format
    ws.cell(row=row, column=6, value=status)
    if status == "LOSS":
        ws.cell(row=row, column=6).fill = red_fill
    elif status == "STRONG":
        ws.cell(row=row, column=6).fill = green_fill
    else:
        ws.cell(row=row, column=6).fill = yellow_fill
    row += 1

# Total
ws.cell(row=row, column=1, value="TOTAL").font = bold_font
for col in [2, 3, 4]:
    ws.cell(row=row, column=col, value=f"=SUM({get_column_letter(col)}{seg_start}:{get_column_letter(col)}{row-1})")
    ws.cell(row=row, column=col).number_format = currency_format
    ws.cell(row=row, column=col).font = bold_font
ws.cell(row=row, column=5, value=f"=SUM(E{seg_start}:E{row-1})")
ws.cell(row=row, column=5).number_format = percent_format
ws.cell(row=row, column=5).font = bold_font

row += 3

# Key Variances requiring attention
ws.cell(row=row, column=1, value="Key Variances Requiring Board Attention").font = Font(bold=True, size=12, color="C0392B")
row += 1

key_variances = [
    "1. Bloom Marketing segment losing $586K YTD despite $261K revenue - hauling COGS issue",
    "2. Events at 24% of Q1 target (3-4 events vs 13 budgeted) - pipeline review needed",
    "3. RECs significantly outperforming - $2.52M YTD vs $1.22M Q1 budget (207%)",
    "4. AR aging: $90K (61%) over 90 days - collections actions approved",
    "5. Interest income strong at $72K (99% of Q1 budget) from $14.3M cash position",
]

for var in key_variances:
    ws.cell(row=row, column=1, value=var)
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 12

# ============================================================================
# Add Variance Chart to Dashboard
# ============================================================================
dashboard = wb["Dashboard"]

# Add variance comparison data for chart
dashboard['A35'] = "YTD vs Q1 Budget Variance"
dashboard['A35'].font = Font(bold=True, size=12)
dashboard['A36'] = "Category"
dashboard['B36'] = "Q1 Budget"
dashboard['C36'] = "YTD Actual"
for cell in ['A36', 'B36', 'C36']:
    dashboard[cell].font = header_font
    dashboard[cell].fill = header_fill

variance_chart_data = [
    ("RECs", 1215163, 2518742),
    ("Bloom Mktg", 826235, 261185),
    ("Events", 174000, 43299),
    ("IP", 85232, 61500),
    ("Interest", 72500, 72048),
]

row = 37
for name, budget, actual in variance_chart_data:
    dashboard.cell(row=row, column=1, value=name)
    dashboard.cell(row=row, column=2, value=budget)
    dashboard.cell(row=row, column=2).number_format = currency_format
    dashboard.cell(row=row, column=3, value=actual)
    dashboard.cell(row=row, column=3).number_format = currency_format
    row += 1

# Create clustered bar chart
bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.style = 10
bar.title = "YTD Actual vs Q1 Budget"
bar.y_axis.title = "Amount ($)"

data = Reference(dashboard, min_col=2, min_row=36, max_col=3, max_row=41)
cats = Reference(dashboard, min_col=1, min_row=37, max_row=41)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.shape = 4
dashboard.add_chart(bar, "D35")

# Save workbook
wb.save("/Users/loveg/Documents/SDO - Workspace/FY26_Revised_Budget_Board_Presentation.xlsx")
print("Added YTD vs Budget sheet and updated Dashboard with variance chart")
print("\nSheets in workbook:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
