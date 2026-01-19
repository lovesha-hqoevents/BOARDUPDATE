from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Color definitions (Blue Drop blue theme)
BLUE_DROP_BLUE = "0077B6"
HEADER_FILL = PatternFill("solid", fgColor=BLUE_DROP_BLUE)
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="CAF0F8")
INPUT_FILL = PatternFill("solid", fgColor="FFFF99")  # Yellow for inputs
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="D5F5E3")
RED_FILL = PatternFill("solid", fgColor="FADBD8")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=BLUE_DROP_BLUE)
SECTION_FONT = Font(bold=True, size=11)
INPUT_FONT = Font(color="0000FF")  # Blue for inputs
FORMULA_FONT = Font(color="000000")  # Black for formulas

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================
# SHEET 1: Budget Summary by Class
# ============================================
ws1 = wb.active
ws1.title = "Budget Summary"

# Title
ws1['A1'] = "Blue Drop, LLC - FY26 Draft Budget"
ws1['A1'].font = TITLE_FONT
ws1['A2'] = "Fiscal Year: October 1, 2025 - September 30, 2026"
ws1['A3'] = "Last Updated:"
ws1['B3'] = "=TODAY()"
ws1['B3'].number_format = "MM/DD/YYYY"

# Classes (columns)
classes = ["Administration", "Bloom Marketing", "Bloom Sales", "Cell Towers",
           "DC Water Innovations", "Events", "RECs", "Store", "Total"]

# Headers row 5
ws1['A5'] = "Category"
ws1['A5'].font = HEADER_FONT
ws1['A5'].fill = HEADER_FILL
for i, cls in enumerate(classes, start=2):
    cell = ws1.cell(row=5, column=i, value=cls)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

# Revenue section
revenue_items = [
    ("REVENUE", None, True),
    ("Bloom Sales", [0, 0, 425000, 0, 0, 0, 0, 0]),
    ("Bloom Deliveries", [0, 0, 125000, 0, 0, 0, 0, 0]),
    ("Bloom Hauling - DC Water", [0, 1501808, 0, 0, 0, 0, 0, 0]),
    ("Bloom Marketing Fee", [0, 1054130, 0, 0, 0, 0, 0, 0]),
    ("Land App/Storage Reimbursement", [0, 749000, 0, 0, 0, 0, 0, 0]),
    ("Products and IP Revenue", [0, 0, 0, 0, 340929, 0, 0, 0]),
    ("Event Rental Revenue", [0, 0, 0, 0, 0, 696000, 0, 0]),
    ("REC Sales", [0, 0, 0, 0, 0, 0, 4360650, 0]),
    ("Solar REC Sales (NEW)", [0, 0, 0, 0, 0, 0, 500000, 0]),
    ("REC Admin Fee", [0, 0, 0, 0, 0, 0, 50000, 0]),
    ("Cell Tower Revenue", [0, 0, 0, 179972, 0, 0, 0, 0]),
    ("Demand Response Payments", [10000, 0, 0, 0, 0, 0, 0, 0]),
    ("Web Store/Merchandise", [0, 0, 0, 0, 0, 0, 0, 2750]),
    ("Services Revenue", [0, 48000, 0, 0, 0, 0, 0, 0]),
    ("Interest Income", [350000, 0, 0, 0, 0, 0, 0, 0]),
    ("TOTAL REVENUE", None, True),
]

row = 7
revenue_start = row + 1
for item in revenue_items:
    name = item[0]
    ws1.cell(row=row, column=1, value=name)

    if item[1] is None:  # Section header or total
        ws1.cell(row=row, column=1).font = SECTION_FONT
        if "TOTAL" in name:
            # Add SUM formulas for total row
            for col in range(2, 10):
                if col < 10:  # Not the Total column
                    col_letter = get_column_letter(col)
                    ws1.cell(row=row, column=col, value=f"=SUM({col_letter}{revenue_start}:{col_letter}{row-1})")
                    ws1.cell(row=row, column=col).font = SECTION_FONT
                    ws1.cell(row=row, column=col).number_format = '"$"#,##0'
            # Total column sums across
            ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
            ws1.cell(row=row, column=10).font = SECTION_FONT
            ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    else:
        values = item[1]
        for col, val in enumerate(values, start=2):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.number_format = '"$"#,##0'
            if val > 0:
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
        # Total column
        ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
        ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    row += 1

revenue_total_row = row - 1

# COGS Section
row += 1
cogs_items = [
    ("COST OF GOODS SOLD", None, True),
    ("Bloom Hauling COGS", [0, 1400000, 0, 0, 0, 0, 0, 0]),
    ("Blending Material", [0, 0, 280000, 0, 0, 0, 0, 0]),
    ("Subcontractors - Events", [0, 0, 0, 0, 0, 40000, 0, 0]),
    ("Subcontractors - RECs", [0, 0, 0, 0, 0, 0, 50000, 0]),
    ("TOTAL COGS", None, True),
]

cogs_start = row + 1
for item in cogs_items:
    name = item[0]
    ws1.cell(row=row, column=1, value=name)

    if item[1] is None:
        ws1.cell(row=row, column=1).font = SECTION_FONT
        if "TOTAL" in name:
            for col in range(2, 10):
                col_letter = get_column_letter(col)
                ws1.cell(row=row, column=col, value=f"=SUM({col_letter}{cogs_start}:{col_letter}{row-1})")
                ws1.cell(row=row, column=col).font = SECTION_FONT
                ws1.cell(row=row, column=col).number_format = '"$"#,##0'
            ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
            ws1.cell(row=row, column=10).font = SECTION_FONT
            ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    else:
        values = item[1]
        for col, val in enumerate(values, start=2):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.number_format = '"$"#,##0'
            if val > 0:
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
        ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
        ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    row += 1

cogs_total_row = row - 1

# Gross Profit
row += 1
ws1.cell(row=row, column=1, value="GROSS PROFIT")
ws1.cell(row=row, column=1).font = SECTION_FONT
for col in range(2, 11):
    col_letter = get_column_letter(col)
    ws1.cell(row=row, column=col, value=f"={col_letter}{revenue_total_row}-{col_letter}{cogs_total_row}")
    ws1.cell(row=row, column=col).font = SECTION_FONT
    ws1.cell(row=row, column=col).number_format = '"$"#,##0'
gross_profit_row = row

# Operating Expenses
row += 2
expense_items = [
    ("OPERATING EXPENSES", None, True),
    ("Compensation - Salaries", [280000, 420000, 0, 0, 0, 200000, 0, 0]),
    ("Compensation - Fringe", [70000, 105000, 0, 0, 0, 50000, 0, 0]),
    ("Compensation - Payroll Taxes", [28000, 42000, 0, 0, 0, 20000, 0, 0]),
    ("Compensation - Bonus", [40000, 80000, 0, 0, 0, 20000, 0, 0]),
    ("Professional Fees - Legal", [15000, 0, 0, 0, 0, 0, 0, 0]),
    ("Professional Fees - Audit", [30000, 0, 0, 0, 0, 0, 0, 0]),
    ("Professional Fees - Accounting", [24000, 0, 0, 0, 0, 0, 0, 0]),
    ("Professional Fees - HR", [12000, 8000, 0, 0, 0, 5000, 0, 0]),
    ("Professional Fees - Land App/Storage", [0, 400000, 0, 0, 0, 0, 0, 0]),
    ("Professional Fees - Other", [5000, 30000, 0, 0, 0, 10000, 0, 0]),
    ("Travel & Transportation", [2000, 8000, 0, 0, 0, 2000, 0, 0]),
    ("Technology & Software", [15000, 8000, 0, 0, 0, 5000, 5000, 0]),
    ("Equipment Maintenance", [0, 20000, 0, 0, 0, 5000, 0, 0]),
    ("Bank/Merchant Fees", [0, 15000, 0, 0, 0, 3000, 0, 500]),
    ("Office Supplies", [2000, 2000, 0, 0, 0, 2000, 0, 0]),
    ("Dues & Subscriptions", [40000, 20000, 0, 0, 0, 45000, 5000, 0]),
    ("Training & Conferences", [5000, 10000, 0, 0, 0, 5000, 0, 0]),
    ("Business Meals/Entertainment", [8000, 5000, 0, 0, 0, 3000, 0, 0]),
    ("Depreciation", [35000, 0, 0, 0, 0, 0, 0, 0]),
    ("Administration Other", [15000, 5000, 0, 0, 0, 3000, 0, 0]),
    ("Employee Relations", [10000, 3000, 0, 0, 0, 2000, 0, 0]),
    ("Insurance", [25000, 0, 0, 0, 0, 0, 0, 0]),
    ("Licenses & Permits", [3000, 0, 0, 0, 0, 5000, 0, 0]),
    ("Marketing - Website", [0, 10000, 0, 0, 0, 0, 0, 0]),
    ("Marketing - Print/Promo", [0, 8000, 0, 0, 0, 5000, 0, 0]),
    ("Marketing - Advertising", [0, 15000, 0, 0, 0, 8000, 0, 0]),
    ("Taxes & Licenses", [5000, 0, 0, 0, 0, 0, 0, 0]),
    ("TOTAL OPERATING EXPENSES", None, True),
]

expense_start = row + 1
for item in expense_items:
    name = item[0]
    ws1.cell(row=row, column=1, value=name)

    if item[1] is None:
        ws1.cell(row=row, column=1).font = SECTION_FONT
        if "TOTAL" in name:
            for col in range(2, 10):
                col_letter = get_column_letter(col)
                ws1.cell(row=row, column=col, value=f"=SUM({col_letter}{expense_start}:{col_letter}{row-1})")
                ws1.cell(row=row, column=col).font = SECTION_FONT
                ws1.cell(row=row, column=col).number_format = '"$"#,##0'
            ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
            ws1.cell(row=row, column=10).font = SECTION_FONT
            ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    else:
        values = item[1]
        for col, val in enumerate(values, start=2):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.number_format = '"$"#,##0'
            if val > 0:
                cell.font = INPUT_FONT
                cell.fill = INPUT_FILL
        ws1.cell(row=row, column=10, value=f"=SUM(B{row}:I{row})")
        ws1.cell(row=row, column=10).number_format = '"$"#,##0'
    row += 1

expense_total_row = row - 1

# Net Operating Income
row += 1
ws1.cell(row=row, column=1, value="NET OPERATING INCOME")
ws1.cell(row=row, column=1).font = Font(bold=True, size=12)
for col in range(2, 11):
    col_letter = get_column_letter(col)
    ws1.cell(row=row, column=col, value=f"={col_letter}{gross_profit_row}-{col_letter}{expense_total_row}")
    ws1.cell(row=row, column=col).font = Font(bold=True, size=12)
    ws1.cell(row=row, column=col).number_format = '"$"#,##0'
net_income_row = row

# Set column widths
ws1.column_dimensions['A'].width = 35
for col in range(2, 11):
    ws1.column_dimensions[get_column_letter(col)].width = 18

# ============================================
# SHEET 2: YTD Actuals vs Budget
# ============================================
ws2 = wb.create_sheet("YTD vs Budget")

ws2['A1'] = "Blue Drop, LLC - FY26 YTD Performance vs Budget"
ws2['A1'].font = TITLE_FONT
ws2['A2'] = "Through Date:"
ws2['B2'] = "January 19, 2026"
ws2['A3'] = "Days into FY:"
ws2['B3'] = 111  # Oct 1 to Jan 19
ws2['A4'] = "% of Year:"
ws2['B4'] = "=B3/365"
ws2['B4'].number_format = "0.0%"

# Headers
headers = ["Category", "FY26 Budget", "YTD Budget", "YTD Actual", "Variance $", "Variance %", "Full Year Forecast"]
for i, h in enumerate(headers, start=1):
    cell = ws2.cell(row=6, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

# Data rows with formulas
ytd_data = [
    ("REVENUE", None, None, None),
    ("RECs Revenue", 4910650, 1493085, 2518742),
    ("Bloom Marketing", 3352938, 1019602, 261185),
    ("Bloom Sales", 550000, 167260, 90769),
    ("DC Water IP/Licensing", 340929, 103673, 61500),
    ("Events Revenue", 696000, 211644, 43299),
    ("Cell Towers", 179972, 54739, 43182),
    ("Interest Income", 350000, 106438, 72048),
    ("Other Revenue", 50750, 15436, 2179),
    ("TOTAL REVENUE", None, None, None),
]

row = 7
rev_start = row + 1
for item in ytd_data:
    name, budget, ytd_budget, ytd_actual = item
    ws2.cell(row=row, column=1, value=name)

    if budget is None:
        ws2.cell(row=row, column=1).font = SECTION_FONT
        if "TOTAL" in name:
            for col in range(2, 8):
                col_letter = get_column_letter(col)
                ws2.cell(row=row, column=col, value=f"=SUM({col_letter}{rev_start}:{col_letter}{row-1})")
                ws2.cell(row=row, column=col).font = SECTION_FONT
    else:
        ws2.cell(row=row, column=2, value=budget)
        ws2.cell(row=row, column=2).number_format = '"$"#,##0'
        ws2.cell(row=row, column=2).font = INPUT_FONT
        ws2.cell(row=row, column=2).fill = INPUT_FILL

        # YTD Budget = Full Budget * % of Year
        ws2.cell(row=row, column=3, value=f"=B{row}*$B$4")
        ws2.cell(row=row, column=3).number_format = '"$"#,##0'

        ws2.cell(row=row, column=4, value=ytd_actual)
        ws2.cell(row=row, column=4).number_format = '"$"#,##0'
        ws2.cell(row=row, column=4).font = INPUT_FONT
        ws2.cell(row=row, column=4).fill = INPUT_FILL

        # Variance $ = Actual - Budget
        ws2.cell(row=row, column=5, value=f"=D{row}-C{row}")
        ws2.cell(row=row, column=5).number_format = '"$"#,##0;("$"#,##0)'

        # Variance % = Variance / Budget
        ws2.cell(row=row, column=6, value=f"=IF(C{row}=0,0,E{row}/C{row})")
        ws2.cell(row=row, column=6).number_format = "0.0%"

        # Full Year Forecast = Actual / % of Year
        ws2.cell(row=row, column=7, value=f"=IF($B$4=0,0,D{row}/$B$4)")
        ws2.cell(row=row, column=7).number_format = '"$"#,##0'

    row += 1

# Set column widths
ws2.column_dimensions['A'].width = 30
for col in range(2, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# ============================================
# SHEET 3: Monthly Tracking
# ============================================
ws3 = wb.create_sheet("Monthly Tracking")

ws3['A1'] = "Blue Drop, LLC - FY26 Monthly Budget Tracking"
ws3['A1'].font = TITLE_FONT

months = ["Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "FY Total"]

# Headers
ws3.cell(row=3, column=1, value="Revenue Category").font = HEADER_FONT
ws3.cell(row=3, column=1).fill = HEADER_FILL
for i, m in enumerate(months, start=2):
    cell = ws3.cell(row=3, column=i, value=m)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

# Monthly revenue rows
monthly_items = [
    "RECs Revenue",
    "Bloom Marketing",
    "Bloom Sales",
    "DC Water IP",
    "Events",
    "Cell Towers",
    "Interest Income",
    "Other",
    "TOTAL REVENUE"
]

row = 4
for item in monthly_items:
    ws3.cell(row=row, column=1, value=item)
    if item == "TOTAL REVENUE":
        ws3.cell(row=row, column=1).font = SECTION_FONT
        for col in range(2, 15):
            col_letter = get_column_letter(col)
            ws3.cell(row=row, column=col, value=f"=SUM({col_letter}4:{col_letter}{row-1})")
            ws3.cell(row=row, column=col).font = SECTION_FONT
            ws3.cell(row=row, column=col).number_format = '"$"#,##0'
    else:
        for col in range(2, 14):
            cell = ws3.cell(row=row, column=col, value=0)
            cell.number_format = '"$"#,##0'
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
        # FY Total formula
        row_letter = get_column_letter(14)
        ws3.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        ws3.cell(row=row, column=14).number_format = '"$"#,##0'
    row += 1

ws3.column_dimensions['A'].width = 20
for col in range(2, 15):
    ws3.column_dimensions[get_column_letter(col)].width = 12

# ============================================
# SHEET 4: Budget Review Process
# ============================================
ws4 = wb.create_sheet("Review Process")

ws4['A1'] = "FY26 Budget Review & Update Process"
ws4['A1'].font = TITLE_FONT

process_content = [
    ("", ""),
    ("BUDGET REVIEW SCHEDULE", ""),
    ("", ""),
    ("Frequency", "Activity"),
    ("Weekly", "Update YTD Actuals from QuickBooks P&L by Class report"),
    ("Weekly", "Review AR aging and update collection forecasts"),
    ("Monthly", "Full variance analysis - compare actual vs budget by class"),
    ("Monthly", "Update full-year forecast based on trends"),
    ("Monthly", "Department heads review their class budgets"),
    ("Quarterly", "Board presentation with variance explanations"),
    ("Quarterly", "Revise annual forecast if variance >10%"),
    ("", ""),
    ("DATA SOURCES", ""),
    ("", ""),
    ("Report", "Source"),
    ("P&L by Class", "QuickBooks > Reports > Profit and Loss by Class"),
    ("AR Aging", "QuickBooks > Reports > A/R Aging Summary"),
    ("Cash Position", "QuickBooks > Reports > Balance Sheet"),
    ("REC Revenue", "PJM GATS system / Broker statements"),
    ("Event Revenue", "HubSpot CRM + Invoicing system"),
    ("", ""),
    ("VARIANCE THRESHOLDS", ""),
    ("", ""),
    ("Threshold", "Action Required"),
    (">5% under budget", "Monitor - note in monthly review"),
    (">10% under budget", "Investigate root cause, document findings"),
    (">15% under budget", "Develop corrective action plan"),
    (">20% under budget", "Escalate to leadership, revise forecast"),
    ("", ""),
    ("BUDGET OWNERS BY CLASS", ""),
    ("", ""),
    ("Class", "Owner"),
    ("Administration", "CFO / Controller"),
    ("Bloom Marketing", "VP Operations"),
    ("Bloom Sales", "Sales Manager"),
    ("Cell Towers", "CFO (passive income)"),
    ("DC Water Innovations", "CEO"),
    ("Events", "Events Coordinator (Raul Buddle)"),
    ("RECs", "CFO / Broker liaison"),
    ("Store", "Marketing Manager"),
    ("", ""),
    ("REVISION APPROVAL PROCESS", ""),
    ("", ""),
    ("Change Amount", "Approval Required"),
    ("<$10,000", "Department Head"),
    ("$10,000 - $50,000", "CFO approval"),
    ("$50,000 - $100,000", "CEO approval"),
    (">$100,000", "Board approval"),
    ("", ""),
    ("KEY CONTACTS", ""),
    ("", ""),
    ("Role", "Responsibility"),
    ("Adam Lawrence", "Weekly data entry, AR tracking, invoice processing"),
    ("Raul Buddle", "Events revenue tracking, pipeline updates"),
    ("Controller", "Monthly close, variance analysis, board reporting"),
    ("CEO", "Strategic decisions, major variance approvals"),
]

for i, (col1, col2) in enumerate(process_content, start=3):
    ws4.cell(row=i, column=1, value=col1)
    ws4.cell(row=i, column=2, value=col2)
    if col1 and col1.isupper() and col2 == "":
        ws4.cell(row=i, column=1).font = SECTION_FONT
        ws4.cell(row=i, column=1).fill = LIGHT_BLUE_FILL
    elif col1 in ["Frequency", "Report", "Threshold", "Class", "Change Amount", "Role"]:
        ws4.cell(row=i, column=1).font = Font(bold=True)
        ws4.cell(row=i, column=2).font = Font(bold=True)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 60

# ============================================
# SHEET 5: Assumptions
# ============================================
ws5 = wb.create_sheet("Assumptions")

ws5['A1'] = "FY26 Budget Assumptions"
ws5['A1'].font = TITLE_FONT

assumptions = [
    ("", ""),
    ("REVENUE ASSUMPTIONS", ""),
    ("", ""),
    ("Category", "Assumption", "Source/Basis"),
    ("RECs", "4,360,650 base + 500,000 Solar RECs", "Historical sales + Q1 FY26 Solar performance"),
    ("Bloom Marketing Fee", "Based on DC Water contract terms", "Existing contract"),
    ("Bloom Hauling", "1,501,808 per contract", "DC Water contract"),
    ("Land App Reimbursement", "749,000 annual", "Historical + contracts"),
    ("Events", "52 events @ $13,385 average", "Q1 performance adjusted"),
    ("Cell Towers", "179,972 annual", "Existing lease agreements"),
    ("Interest Income", "350,000 @ 5.5% yield on $6.4M average cash", "Current bank rates"),
    ("", ""),
    ("EXPENSE ASSUMPTIONS", ""),
    ("", ""),
    ("Category", "Assumption", "Source/Basis"),
    ("Salaries", "3% increase from FY25", "Standard COLA"),
    ("Fringe Benefits", "25% of salaries", "Historical ratio"),
    ("Payroll Taxes", "10% of salaries", "FICA + unemployment"),
    ("Hauling COGS", "~93% of Hauling revenue", "Historical margin"),
    ("Land App Fees", "~53% of Land App revenue", "Contract terms"),
    ("Professional Fees", "Based on contracts + estimates", "Vendor quotes"),
    ("", ""),
    ("KEY RISKS", ""),
    ("", ""),
    ("Risk", "Impact", "Mitigation"),
    ("REC price volatility", "Could affect $4.8M+ revenue", "Diversify timing of sales"),
    ("AR collections", "$90K+ at risk (>90 days)", "Collections campaign, new processes"),
    ("Events underperformance", "Only 24% of Q1 target achieved", "Pipeline review, strategy refresh"),
    ("Bloom Marketing losses", "$586K loss YTD", "Cost structure review by Q3"),
]

for i, row_data in enumerate(assumptions, start=3):
    for j, val in enumerate(row_data):
        cell = ws5.cell(row=i, column=j+1, value=val)
        if val and str(val).isupper() and len(row_data) == 2:
            cell.font = SECTION_FONT
            cell.fill = LIGHT_BLUE_FILL
        elif val in ["Category", "Risk"]:
            cell.font = Font(bold=True)

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 45
ws5.column_dimensions['C'].width = 40

# Save workbook
output_path = "/Users/loveg/Documents/SDO - Workspace/FY26_Draft_Budget.xlsx"
wb.save(output_path)
print(f"Budget created: {output_path}")
